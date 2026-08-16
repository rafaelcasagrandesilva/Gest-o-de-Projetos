"""Workspace Jurídico (Fase 1) — permissões, redação de valores e agregados do serviço.

Cobre as três decisões que mais custam caro se regredirem:
1. o módulo é isolado (nenhuma permissão de outro módulo abre o Jurídico, e vice-versa);
2. sem `legal.sensitive` o BACKEND não envia os valores (não é ocultação no frontend);
3. KPIs/gráficos e a lista respondem ao MESMO conjunto de filtros.
"""

from __future__ import annotations

import unittest
from types import SimpleNamespace

from app.api.deps import user_has_permission
from app.api.sensitive import redact_for
from app.core import permission_codes as pc
from app.core.session_context import accessible_workspaces, session_permission_names
from app.schemas.legal import (
    LegalBucket,
    LegalCaseRead,
    LegalFacets,
    LegalKpis,
    LegalOverview,
    LegalPersonDetail,
)
from app.services.legal_service import STATUS_LABELS, STATUS_ORDER, TYPE_LABELS, CaseFilters


def _user(*perms: str, roles: list[str] | None = None) -> SimpleNamespace:
    ups = [SimpleNamespace(permission=SimpleNamespace(name=p), granted=True) for p in perms]
    role_links = []
    for name in roles or []:
        preset = pc.ROLE_PRESET.get(name, frozenset())
        rp = [SimpleNamespace(permission=SimpleNamespace(name=p)) for p in preset]
        role_links.append(SimpleNamespace(role=SimpleNamespace(name=name, permissions=rp)))
    return SimpleNamespace(
        email="legal-test@example.com", user_permissions=ups, roles=role_links, is_active=True
    )


def _case(**overrides) -> LegalCaseRead:
    data = {
        "id": "00000000-0000-0000-0000-000000000001",
        "case_number": "0000001-11.2025.5.24.0086",
        "status": "EM_ANDAMENTO",
        "case_type": "TRABALHISTA",
        "amount_claimed": 1000.0,
        "amount_considered": 900.0,
        "amount_agreed": 800.0,
        "amount_paid": 700.0,
        "amount_pending": 600.0,
        "agreement_terms": "3 X 2.333,34",
        "claimant_name": "Fulano de Tal",
        "uf": "MS",
    }
    data.update(overrides)
    return LegalCaseRead.model_validate(data)


class PermissionIsolationTests(unittest.TestCase):
    """Um recurso por MENU: nenhum menu abre outro, e nada de fora abre o Jurídico."""

    def test_verb_chain_per_menu(self):
        # update ⇒ read ⇒ list ⇒ reference, dentro do MESMO recurso.
        u = _user(pc.LEGAL_CASES_UPDATE)
        for code in (pc.LEGAL_CASES_READ, pc.LEGAL_CASES_LIST, pc.LEGAL_CASES_REFERENCE):
            self.assertTrue(user_has_permission(u, code), code)
        self.assertFalse(user_has_permission(u, pc.LEGAL_CASES_SENSITIVE))
        self.assertFalse(user_has_permission(u, pc.LEGAL_CASES_CREATE))

    def test_menus_do_not_leak_into_each_other(self):
        """O ponto central da granularidade: editar Processos não toca em Desligados."""
        cases_admin = _user(
            pc.LEGAL_CASES_CREATE, pc.LEGAL_CASES_UPDATE, pc.LEGAL_CASES_DELETE,
            pc.LEGAL_CASES_SENSITIVE,
        )
        for code in (
            pc.LEGAL_PERSONS_LIST, pc.LEGAL_PERSONS_READ, pc.LEGAL_PERSONS_CREATE,
            pc.LEGAL_PERSONS_UPDATE, pc.LEGAL_PERSONS_DELETE, pc.LEGAL_PERSONS_SENSITIVE,
            pc.LEGAL_COMPANIES_UPDATE, pc.LEGAL_PROJECTS_UPDATE, pc.LEGAL_REPORTS_EXPORT,
        ):
            self.assertFalse(user_has_permission(cases_admin, code), f"Processos não pode dar {code}")

        persons_admin = _user(pc.LEGAL_PERSONS_UPDATE, pc.LEGAL_PERSONS_SENSITIVE)
        for code in (pc.LEGAL_CASES_UPDATE, pc.LEGAL_CASES_SENSITIVE, pc.LEGAL_CASES_DELETE):
            self.assertFalse(user_has_permission(persons_admin, code), f"Desligados não pode dar {code}")

    def test_sensitive_is_independent_per_menu(self):
        self.assertFalse(user_has_permission(_user(pc.LEGAL_CASES_READ), pc.LEGAL_CASES_SENSITIVE))
        # Ter os valores de Processos NÃO dá os valores de Desligados.
        u = _user(pc.LEGAL_CASES_SENSITIVE)
        self.assertTrue(user_has_permission(u, pc.LEGAL_CASES_SENSITIVE))
        self.assertFalse(user_has_permission(u, pc.LEGAL_PERSONS_SENSITIVE))

    def test_other_modules_do_not_open_legal(self):
        u = _user(
            pc.EMPLOYEES_VIEW, pc.PROJECTS_VIEW, pc.PAYABLES_VIEW, pc.SYSTEM_ADMIN, pc.ASSETS_EDIT,
            pc.REPORTS_VIEW, pc.REPORTS_EXPORT,
        )
        for code in (*pc.LEGAL_MODULE_CODES, pc.WORKSPACE_LEGAL_ACCESS):
            self.assertFalse(user_has_permission(u, code), f"{code} não pode vir de outro módulo")

    def test_legal_does_not_open_other_modules(self):
        u = _user(*pc.LEGAL_MODULE_CODES)
        for code in (
            pc.EMPLOYEES_READ, pc.PROJECTS_READ, pc.PAYABLES_READ, pc.USERS_MANAGE,
            pc.WORKSPACE_PROJECTS_ACCESS, pc.WORKSPACE_FINANCE_ACCESS,
            # Exportar o Jurídico NÃO concede o módulo de Relatórios corporativo.
            pc.REPORTS_READ, pc.REPORTS_EXPORT,
        ):
            self.assertFalse(user_has_permission(u, code), f"o Jurídico não pode liberar {code}")

    def test_workspace_access_derives_from_any_menu(self):
        for code in (
            pc.LEGAL_DASHBOARD_READ, pc.LEGAL_CASES_LIST, pc.LEGAL_PERSONS_LIST,
            pc.LEGAL_COMPANIES_LIST, pc.LEGAL_PROJECTS_LIST, pc.LEGAL_REPORTS_READ,
        ):
            self.assertTrue(
                user_has_permission(_user(code), pc.WORKSPACE_LEGAL_ACCESS), f"{code} deve abrir"
            )
        # `reference` e `sensitive` sozinhos NÃO abrem o workspace.
        for code in (pc.LEGAL_CASES_REFERENCE, pc.LEGAL_CASES_SENSITIVE):
            self.assertFalse(
                user_has_permission(_user(code), pc.WORKSPACE_LEGAL_ACCESS), f"{code} não abre"
            )
        self.assertIn("legal", accessible_workspaces(_user(pc.LEGAL_CASES_READ)))
        self.assertNotIn("legal", accessible_workspaces(_user(pc.EMPLOYEES_VIEW)))

    def test_codes_are_active_in_session(self):
        names = session_permission_names(_user(pc.LEGAL_CASES_READ, pc.LEGAL_CASES_SENSITIVE))
        self.assertIn(pc.LEGAL_CASES_READ, names)
        self.assertIn(pc.LEGAL_CASES_SENSITIVE, names)
        self.assertIn(pc.WORKSPACE_LEGAL_ACCESS, names)

    def test_role_presets(self):
        gestor, consulta = _user(roles=["GESTOR"]), _user(roles=["CONSULTA"])
        # GESTOR administra o módulo inteiro.
        for code in pc.LEGAL_MODULE_CODES:
            self.assertTrue(user_has_permission(gestor, code), f"GESTOR deveria ter {code}")
        # CONSULTA lê todos os menus, sem valores e sem CRUD.
        for code in (
            pc.LEGAL_DASHBOARD_READ, pc.LEGAL_CASES_READ, pc.LEGAL_PERSONS_READ,
            pc.LEGAL_COMPANIES_LIST, pc.LEGAL_PROJECTS_LIST, pc.LEGAL_REPORTS_READ,
        ):
            self.assertTrue(user_has_permission(consulta, code), f"CONSULTA deveria ter {code}")
        for code in (
            pc.LEGAL_CASES_SENSITIVE, pc.LEGAL_PERSONS_SENSITIVE, pc.LEGAL_CASES_UPDATE,
            pc.LEGAL_PERSONS_DELETE, pc.LEGAL_COMPANIES_CREATE, pc.LEGAL_REPORTS_EXPORT,
        ):
            self.assertFalse(user_has_permission(consulta, code), f"CONSULTA NÃO pode ter {code}")

    def test_every_module_code_is_active_and_catalogued(self):
        from app.core.permission_codes import ACTIVE_PERMISSION_CODES, ALL_PERMISSION_CODES

        specs = {s.code for s in pc.PERMISSION_SPECS}
        for code in pc.LEGAL_MODULE_CODES:
            self.assertIn(code, ALL_PERMISSION_CODES, code)
            self.assertIn(code, ACTIVE_PERMISSION_CODES, code)
            self.assertIn(code, specs, f"{code} sem descrição na grade de administração")


class RedactionTests(unittest.TestCase):
    """Sem `legal.sensitive` o payload sai SEM os valores; o resto da informação permanece."""

    def test_case_values_omitted(self):
        without = redact_for("legal_case", _case(), _user(pc.LEGAL_CASES_READ))
        for field in (
            "amount_claimed",
            "amount_considered",
            "amount_agreed",
            "amount_paid",
            "amount_pending",
            "agreement_terms",
        ):
            self.assertIsNone(getattr(without, field), f"{field} deveria ser omitido")
        # Estrutura preservada: número, status, partes e UF continuam.
        self.assertEqual(without.case_number, "0000001-11.2025.5.24.0086")
        self.assertEqual(without.status.value, "EM_ANDAMENTO")
        self.assertEqual(without.claimant_name, "Fulano de Tal")
        self.assertEqual(without.uf, "MS")

    def test_case_values_kept_with_permission(self):
        with_perm = redact_for("legal_case", _case(), _user(pc.LEGAL_CASES_READ, pc.LEGAL_CASES_SENSITIVE))
        self.assertEqual(with_perm.amount_considered, 900.0)
        self.assertEqual(with_perm.agreement_terms, "3 X 2.333,34")

    def test_person_redaction_covers_only_its_own_fields(self):
        person = LegalPersonDetail.model_validate(
            {
                "id": "00000000-0000-0000-0000-0000000000aa",
                "full_name": "Fulano de Tal",
                "severance_amount": 5000.0,
                "fgts_balance": 400.0,
                "case_count": 1,
                "total_claimed": 1000.0,
                "total_considered": 900.0,
                "cases": [_case()],
            }
        )
        without = redact_for("legal_person", person, _user(pc.LEGAL_CASES_READ))
        self.assertIsNone(without.severance_amount)
        self.assertIsNone(without.fgts_balance)
        self.assertIsNone(without.total_considered)
        # Os processos ANINHADOS são de OUTRO recurso: este redator não os toca de propósito
        # (propagar a decisão do pai vazava valores — ver NestedRedactionRegressionTests). Quem
        # compõe os dois níveis é o router, cada um com o seu gate.
        self.assertEqual(without.cases[0].amount_considered, 900.0)
        # A contagem NÃO é valor monetário e permanece.
        self.assertEqual(without.case_count, 1)

    def test_overview_redaction_is_recursive(self):
        overview = LegalOverview(
            kpis=LegalKpis(case_count=3, person_count=2, total_considered=900.0, total_claimed=1000.0),
            by_status=[LegalBucket(key="EM_ANDAMENTO", label="Em andamento", value=900.0, count=3)],
            by_uf=[LegalBucket(key="MS", label="MS", value=900.0, count=3)],
            facets=LegalFacets(),
        )
        without = redact_for("legal_overview", overview, _user(pc.LEGAL_CASES_LIST))
        self.assertIsNone(without.kpis.total_considered)
        self.assertIsNone(without.by_status[0].value)
        self.assertIsNone(without.by_uf[0].value)
        # Contagens permanecem — os gráficos continuam informativos sem expor o passivo.
        self.assertEqual(without.kpis.case_count, 3)
        self.assertEqual(without.by_status[0].count, 3)


class FilterAndLabelTests(unittest.TestCase):
    def test_basis_selects_the_value_column(self):
        self.assertEqual(CaseFilters().basis_column().key, "amount_considered")
        self.assertEqual(CaseFilters(basis="claimed").basis_column().key, "amount_claimed")

    def test_labels_cover_every_enum_value(self):
        from app.models.legal import LegalCaseStatus, LegalCaseType

        for status in LegalCaseStatus:
            self.assertIn(status.value, STATUS_LABELS)
            self.assertIn(status.value, STATUS_ORDER)
        for case_type in LegalCaseType:
            self.assertIn(case_type.value, TYPE_LABELS)

    def test_default_filters_exclude_inactive(self):
        """Fase 2: sem filtro nenhum, a única condição é `is_active` — processo desativado não
        pode aparecer nas telas analíticas nem entrar nos indicadores."""
        from app.services.legal_service import LegalService

        base = LegalService._case_conditions(CaseFilters())
        self.assertEqual(len(base), 1)
        self.assertIn("is_active", str(base[0]))
        # Só a Administração pede o acervo completo.
        self.assertEqual(LegalService._case_conditions(CaseFilters(include_inactive=True)), [])

    def test_each_axis_adds_one_condition(self):
        from app.services.legal_service import LegalService

        def extra(filters: CaseFilters) -> int:
            # Desconta a condição implícita de `is_active`.
            return len(LegalService._case_conditions(filters)) - 1

        self.assertEqual(extra(CaseFilters(ufs=["MS", "MT"])), 1)
        self.assertEqual(extra(CaseFilters(value_min=10, value_max=20)), 2)
        self.assertEqual(extra(CaseFilters(statuses=["ACORDO"], companies=["X"])), 2)


class ChangeLogTests(unittest.TestCase):
    """Fase 2 — histórico: um registro por CAMPO, com valores legíveis."""

    def _svc(self):
        from app.services.legal_service import LegalService

        svc = LegalService.__new__(LegalService)
        svc.session = SimpleNamespace(added=[], add=lambda o: svc.session.added.append(o))
        return svc

    def test_diff_logs_only_what_changed(self):
        from app.models.legal import LegalCaseStatus, LegalEntityType

        svc = self._svc()
        row = SimpleNamespace(
            id="00000000-0000-0000-0000-0000000000cc",
            status=LegalCaseStatus.EM_ANDAMENTO,
            company="Antiga",
            notes=None,
        )
        changed = svc._log_diff(
            entity_type=LegalEntityType.CASE,
            row=row,
            # `company` vem igual: não pode virar histórico nem ser reaplicado.
            data={"status": LegalCaseStatus.ACORDO, "company": "Antiga", "notes": "obs"},
            actor=SimpleNamespace(id=None, email="quem@ex.com"),
        )
        self.assertEqual(set(changed), {"status", "notes"})
        self.assertEqual(len(svc.session.added), 2)
        by_field = {e.field: e for e in svc.session.added}
        # str-Enum precisa sair como "EM_ANDAMENTO", não "LegalCaseStatus.EM_ANDAMENTO".
        self.assertEqual(by_field["status"].old_value, "EM_ANDAMENTO")
        self.assertEqual(by_field["status"].new_value, "ACORDO")
        self.assertEqual(by_field["notes"].old_value, None)
        self.assertEqual(by_field["notes"].new_value, "obs")
        self.assertEqual(by_field["status"].changed_by_email, "quem@ex.com")

    def test_money_fields_are_marked_sensitive(self):
        from app.services.legal_service import MONEY_FIELDS

        for f in ("amount_claimed", "amount_considered", "amount_agreed", "severance_amount"):
            self.assertIn(f, MONEY_FIELDS)
        # Campos estruturais NÃO são redigidos no histórico.
        for f in ("status", "company", "project", "notes"):
            self.assertNotIn(f, MONEY_FIELDS)

    def test_change_log_redacts_money_without_permission(self):
        from app.models.legal import LegalChangeAction, LegalEntityType
        from app.modules.legal.router import _change_log_read

        entry = SimpleNamespace(
            id="00000000-0000-0000-0000-0000000000dd",
            created_at=SimpleNamespace(),
            entity_type=LegalEntityType.CASE,
            entity_id="00000000-0000-0000-0000-0000000000cc",
            action=LegalChangeAction.UPDATE,
            field="amount_considered",
            old_value="1000.00",
            new_value="2000.00",
            changed_by_email="quem@ex.com",
        )
        from app.schemas.legal import LegalChangeLogRead

        # Substitui o created_at por um datetime real para o schema validar.
        from datetime import datetime, timezone

        entry.created_at = datetime.now(timezone.utc)

        allowed = _change_log_read(entry, money_by_entity={"CASE": True})
        self.assertEqual(allowed.new_value, "2000.00")

        denied = _change_log_read(entry, money_by_entity={"CASE": False})
        self.assertIsNone(denied.old_value)
        self.assertIsNone(denied.new_value)
        # O REGISTRO da alteração continua visível — só o valor some.
        self.assertEqual(denied.field, "amount_considered")
        self.assertEqual(denied.action, LegalChangeAction.UPDATE)
        self.assertIsInstance(allowed, LegalChangeLogRead)

    def test_non_money_field_never_redacted(self):
        from datetime import datetime, timezone

        from app.models.legal import LegalChangeAction, LegalEntityType
        from app.modules.legal.router import _change_log_read

        entry = SimpleNamespace(
            id="00000000-0000-0000-0000-0000000000ee",
            created_at=datetime.now(timezone.utc),
            entity_type=LegalEntityType.CASE,
            entity_id="00000000-0000-0000-0000-0000000000cc",
            action=LegalChangeAction.UPDATE,
            field="status",
            old_value="EM_ANDAMENTO",
            new_value="ACORDO",
            changed_by_email=None,
        )
        denied = _change_log_read(entry, money_by_entity={"CASE": False})
        self.assertEqual(denied.old_value, "EM_ANDAMENTO")
        self.assertEqual(denied.new_value, "ACORDO")


    def test_history_money_gate_follows_the_entity(self):
        """Ter `sensitive` de Processos não revela valor de Desligados no histórico."""
        from datetime import datetime, timezone

        from app.models.legal import LegalChangeAction, LegalEntityType
        from app.modules.legal.router import _change_log_read

        def entry(entity, field, value):
            return SimpleNamespace(
                id="00000000-0000-0000-0000-0000000000ff",
                created_at=datetime.now(timezone.utc),
                entity_type=entity,
                entity_id="00000000-0000-0000-0000-0000000000cc",
                action=LegalChangeAction.UPDATE,
                field=field,
                old_value="1.00",
                new_value=value,
                changed_by_email=None,
            )

        gate = {"CASE": True, "PERSON": False}
        caso = _change_log_read(entry(LegalEntityType.CASE, "amount_considered", "9.00"), money_by_entity=gate)
        pessoa = _change_log_read(entry(LegalEntityType.PERSON, "severance_amount", "9.00"), money_by_entity=gate)
        self.assertEqual(caso.new_value, "9.00")
        self.assertIsNone(pessoa.new_value)


def _cell(ws, row: int, header: str):
    """Valor da célula pela COLUNA NOMEADA (o teste não deve depender da ordem das colunas)."""
    headers = [c.value for c in ws[1]]
    return ws.cell(row=row, column=headers.index(header) + 1).value


class ReportTests(unittest.TestCase):
    """Relatório: uma aba por menu, valores respeitando o `sensitive` de CADA recurso."""

    def _payload(self, *, cases_sensitive=True, persons_sensitive=True):
        from app.services.legal_report_service import _money

        return {
            "resumo": [{"indicador": "Passivo considerado", "quantidade": None,
                        "valor": _money(100.0, include=cases_sensitive)}],
            "quebras": [{"grupo": "Status", "item": "Em andamento", "quantidade": 2,
                         "valor": _money(100.0, include=cases_sensitive)}],
            "processos": [{"processo": "1", "valor_considerado": _money(100.0, include=cases_sensitive)}],
            "desligados": [{"nome": "X", "qtd_processos": 1,
                            "rescisao": _money(50.0, include=persons_sensitive)}],
        }

    def test_xlsx_has_one_sheet_per_menu(self):
        import io

        import openpyxl

        from app.services.legal_report_export import render_legal_report_bytes

        raw, name, mime = render_legal_report_bytes("legal", self._payload(), "xlsx", None)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        self.assertEqual(wb.sheetnames, ["Resumo", "Quebras", "Processos", "Desligados"])
        self.assertTrue(name.endswith(".xlsx"))
        self.assertIn("spreadsheetml", mime)

    def test_pdf_is_generated(self):
        from app.services.legal_report_export import render_legal_report_bytes

        raw, name, mime = render_legal_report_bytes("legal", self._payload(), "pdf", None)
        self.assertEqual(raw[:4], b"%PDF")
        self.assertTrue(name.endswith(".pdf"))
        self.assertEqual(mime, "application/pdf")

    def test_money_cells_are_empty_when_redacted(self):
        import io

        import openpyxl

        from app.services.legal_report_export import render_legal_report_bytes

        payload = self._payload(cases_sensitive=False, persons_sensitive=False)
        raw, _, _ = render_legal_report_bytes("legal", payload, "xlsx", None)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        # Célula VAZIA (openpyxl lê "" como None), nunca 0 — "não pode ver" ≠ "é zero".
        for value in (_cell(wb["Processos"], 2, "Valor considerado"), _cell(wb["Resumo"], 2, "Valor")):
            self.assertIn(value, (None, ""))
            self.assertNotIsInstance(value, (int, float))

    def test_each_resource_sensitive_is_independent(self):
        """Ver valores de Processos não revela a rescisão do Desligado (e vice-versa)."""
        import io

        import openpyxl

        from app.services.legal_report_export import render_legal_report_bytes

        payload = self._payload(cases_sensitive=True, persons_sensitive=False)
        raw, _, _ = render_legal_report_bytes("legal", payload, "xlsx", None)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        self.assertIn(_cell(wb["Desligados"], 2, "Rescisão"), (None, ""))
        self.assertEqual(_cell(wb["Processos"], 2, "Valor considerado"), 100.0)

    def test_report_type_is_registered_everywhere(self):
        from app.schemas.reports import ReportGenerateRequest
        from app.services.export.report_meta import REPORT_TITLES

        self.assertIn("legal", REPORT_TITLES)
        body = ReportGenerateRequest(type="legal", format="xlsx", filters={})
        self.assertEqual(body.type, "legal")


class NestedRedactionRegressionTests(unittest.TestCase):
    """Regressão de um VAZAMENTO real encontrado em auditoria.

    `SensitiveSpec.nested` propaga a decisão do PAI para os filhos (ver `_redact_model`), o que só é
    correto quando pai e filho compartilham a MESMA permissão. A ficha do Desligado cruza recursos
    (`legal_persons.sensitive` × `legal_cases.sensitive`); com `nested`, quem tinha só o sensitive de
    Desligados VIA os valores dos processos aninhados. O router passou a redigir cada nível com o
    seu recurso — estes testes travam esse comportamento.
    """

    def _detail(self):
        return LegalPersonDetail.model_validate(
            {
                "id": "00000000-0000-0000-0000-0000000000aa",
                "full_name": "Fulano",
                "severance_amount": 5000.0,
                "fgts_balance": 400.0,
                "case_count": 1,
                "total_considered": 900.0,
                "cases": [_case()],
            }
        )

    def _render(self, user):
        """Reproduz exatamente o que o router monta em GET /legal/persons/{id}."""
        detail = self._detail()
        with_cases = detail.model_copy(
            update={"cases": [redact_for("legal_case", c, user) for c in detail.cases]}
        )
        return redact_for("legal_person", with_cases, user)

    def test_person_sensitive_alone_does_not_leak_case_values(self):
        out = self._render(_user(pc.LEGAL_PERSONS_READ, pc.LEGAL_PERSONS_SENSITIVE))
        self.assertEqual(out.severance_amount, 5000.0)          # é o recurso dele
        self.assertIsNone(out.cases[0].amount_considered)       # NÃO é — não pode vazar
        self.assertIsNone(out.cases[0].amount_claimed)

    def test_case_sensitive_alone_does_not_leak_person_values(self):
        out = self._render(_user(pc.LEGAL_PERSONS_READ, pc.LEGAL_CASES_SENSITIVE))
        self.assertEqual(out.cases[0].amount_considered, 900.0)
        self.assertIsNone(out.severance_amount)
        self.assertIsNone(out.fgts_balance)

    def test_both_or_neither(self):
        both = self._render(_user(pc.LEGAL_PERSONS_SENSITIVE, pc.LEGAL_CASES_SENSITIVE))
        self.assertEqual(both.severance_amount, 5000.0)
        self.assertEqual(both.cases[0].amount_considered, 900.0)
        none = self._render(_user(pc.LEGAL_PERSONS_READ))
        self.assertIsNone(none.severance_amount)
        self.assertIsNone(none.cases[0].amount_considered)
        self.assertEqual(none.case_count, 1)  # estrutura preservada

    def test_person_spec_has_no_cross_resource_nested(self):
        """Guarda: recolocar `nested` aqui reintroduziria o vazamento."""
        from app.api.sensitive import SENSITIVE_SPECS

        self.assertEqual(SENSITIVE_SPECS["legal_person"].nested, ())


class RestoreIsNotOneWayTests(unittest.TestCase):
    """Quem pode DESATIVAR precisa poder DESFAZER — senão `delete` vira porta de mão única."""

    def test_delete_alone_can_restore(self):
        for delete_code, update_code in (
            (pc.LEGAL_CASES_DELETE, pc.LEGAL_CASES_UPDATE),
            (pc.LEGAL_PERSONS_DELETE, pc.LEGAL_PERSONS_UPDATE),
            (pc.LEGAL_COMPANIES_DELETE, pc.LEGAL_COMPANIES_UPDATE),
            (pc.LEGAL_PROJECTS_DELETE, pc.LEGAL_PROJECTS_UPDATE),
        ):
            u = _user(delete_code)
            self.assertTrue(user_has_permission(u, delete_code))
            # Não ganha update — o endpoint de restaurar é que aceita delete OU update.
            self.assertFalse(user_has_permission(u, update_code), update_code)

    def test_restore_endpoints_accept_delete_or_update(self):
        import inspect

        from app.modules.legal import router as R

        for name, delete_code, update_code in (
            ("_cases_restore", pc.LEGAL_CASES_DELETE, pc.LEGAL_CASES_UPDATE),
            ("_persons_restore", pc.LEGAL_PERSONS_DELETE, pc.LEGAL_PERSONS_UPDATE),
            ("_companies_restore", pc.LEGAL_COMPANIES_DELETE, pc.LEGAL_COMPANIES_UPDATE),
            ("_projects_restore", pc.LEGAL_PROJECTS_DELETE, pc.LEGAL_PROJECTS_UPDATE),
        ):
            dep = getattr(R, name)[0].dependency
            codes = set(inspect.getclosurevars(dep).nonlocals.get("codes", ()))
            self.assertEqual(codes, {delete_code, update_code}, name)


if __name__ == "__main__":
    unittest.main()
