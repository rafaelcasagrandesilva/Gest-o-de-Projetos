"""Relatório Folha de Pagamento (novo): consolidação por colaborador + Excel/PDF.

Não altera o relatório de Colaboradores nem cálculos existentes — apenas lê os dados.
"""

from __future__ import annotations

import io
import unittest
from datetime import date
from uuid import uuid4

from app.services.report_service import _payroll_distribution_label


class _Slice:
    def __init__(self, name, pct):
        self.project_name = name
        self.allocation_percentage = pct


class _Line:
    def __init__(self, slices, admin=0.0):
        self.by_project = slices
        self.administrative_cost = admin


class PayrollHelperTests(unittest.TestCase):
    def test_distribution_labels(self) -> None:
        self.assertEqual(_payroll_distribution_label(None), "—")
        self.assertEqual(_payroll_distribution_label(_Line([_Slice("Fiscalização AT", 100)])), "Fiscalização AT (100%)")
        self.assertEqual(
            _payroll_distribution_label(_Line([_Slice("Fiscalização AT", 50)])),
            "Fiscalização AT (50%) / Administrativo (50%)",
        )
        self.assertEqual(_payroll_distribution_label(_Line([], admin=1200.0)), "Administrativo")
        self.assertEqual(_payroll_distribution_label(_Line([], admin=0.0)), "—")


class PayrollReportDBTests(unittest.IsolatedAsyncioTestCase):
    async def test_consolidation_and_exports(self) -> None:
        from datetime import date as _date

        from sqlalchemy import text
        from sqlalchemy.exc import ProgrammingError

        from app.database.session import AsyncSessionLocal, engine
        from app.models.company_finance import CompanyFinancialItem
        from app.models.employee import Employee
        from app.models.payable_snapshot import PayableSnapshot, PayableSnapshotType
        from app.services.export.report_meta import ReportContext
        from app.services.report_export import render_report_bytes
        from app.services.report_service import ReportService

        await engine.dispose()
        comp = date(2099, 6, 1)          # competência trabalhada (rótulo)
        pay = date(2099, 7, 1)           # CAP que paga esta folha (mês seguinte)
        tag = uuid4().hex[:6]

        async with AsyncSessionLocal() as session:
            try:
                await session.execute(text("SELECT 1 FROM payable_snapshots LIMIT 1"))
            except ProgrammingError:
                self.skipTest("Tabelas ausentes (rode alembic upgrade head).")

            emp = Employee(
                full_name=f"Folha CLT {tag}",
                email=f"folha_{tag}@ex.com",
                employment_type="CLT",
                is_active=True,
                salary_base=5000.0,
                total_cost=0,
                pix_key_type="CPF",
                pix_key="000.000.000-00",
            )
            session.add(emp)
            await session.flush()

            # Endividamento vinculado (item) + seu lançamento no CAP (tipo Endividamento).
            debt = CompanyFinancialItem(
                tipo="endividamento",
                nome=f"Adiantamento {tag}",
                valor_referencia=300.0,
                employee_id=emp.id,
                is_active=True,
                start_date=date(2099, 1, 1),
            )
            session.add(debt)
            await session.flush()

            def _snap(*, type_, ref_id, name, category, value):
                return PayableSnapshot(
                    month=pay,
                    type=type_,
                    ref_id=ref_id,
                    name=name,
                    cost_center="Projeto X",
                    category=category,
                    amount_original=value,
                    amount_final=value,
                    amount_paid=0,
                    due_date=_date(2099, 7, 10),
                    paid=False,
                )

            # Folha do colaborador NO CAP (o relatório deve refletir EXATAMENTE isto):
            session.add(_snap(  # COLLABORATOR — salário
                type_=PayableSnapshotType.COLLABORATOR, ref_id=emp.id,
                name=f"Folha CLT {tag} — Salário CLT", category="Mão de obra", value=4200.00,
            ))
            session.add(_snap(  # COLLABORATOR — benefício
                type_=PayableSnapshotType.COLLABORATOR, ref_id=emp.id,
                name=f"Folha CLT {tag} — Benefício CLT", category="Mão de obra", value=600.00,
            ))
            session.add(_snap(  # ENDIVIDAMENTO vinculado ao colaborador (via ref_id=item)
                type_=PayableSnapshotType.ENDIVIDAMENTO, ref_id=debt.id,
                name=f"Folha CLT {tag}", category="Endividamento", value=300.00,
            ))
            # RUÍDO que NÃO deve entrar (mesma regra da tela): MANUAL e FIXED_COST/Custo diverso.
            session.add(_snap(
                type_=PayableSnapshotType.MANUAL, ref_id=None,
                name=f"Folha CLT {tag} (bônus manual)", category="Manual", value=9999.00,
            ))
            await session.flush()

            svc = ReportService(session)
            payload = await svc.generate_payroll_report(competencia=comp, scenario="REALIZADO")
            rows = payload["rows"]

            mine = [r for r in rows if r["nome"] == f"Folha CLT {tag}"]
            self.assertEqual(len(mine), 1, "deve haver exatamente 1 linha por colaborador")
            r = mine[0]
            self.assertEqual(r["tipo"], "CLT")
            # Componentes = os próprios lançamentos COLLABORATOR do CAP (colunas dinâmicas).
            comps = r["componentes"]
            self.assertEqual(comps.get("Salário CLT"), 4200.00)
            self.assertEqual(comps.get("Benefício CLT"), 600.00)
            self.assertEqual(r["endividamentos"], 300.00)
            # Invariante: total = Σ dos lançamentos do CAP do colaborador (folha + endividamento).
            self.assertEqual(r["total_folha"], 4200.00 + 600.00 + 300.00)
            # O MANUAL (9999) NÃO entra — exatamente como na tela do CAP.
            self.assertNotIn(9999.00, [v for v in comps.values() if v is not None])
            self.assertEqual(r["pix_tipo"], "CPF")

            self.assertIn("Salário CLT", payload["component_columns"])
            self.assertIn("Benefício CLT", payload["component_columns"])

            # Exportações válidas (Excel/PDF).
            ctx = ReportContext(report_type="payroll", generated_by="Teste", filters={"competencia": "2099-06"})
            xlsx, xname, _ = render_report_bytes("payroll", payload, "xlsx", ctx)
            pdf, pname, _ = render_report_bytes("payroll", payload, "pdf", ctx)
            self.assertTrue(xlsx[:2] == b"PK", "xlsx inválido")
            self.assertTrue(pdf[:4] == b"%PDF", "pdf inválido")
            self.assertIn("Folha de Pagamento", xname)

            # Excel deve conter a linha de TOTAIS.
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(xlsx))
            ws = wb.active
            col_a = [str(c.value) for c in ws["A"] if c.value is not None]
            self.assertIn("TOTAIS", col_a)

            # Limpeza (snapshots do CAP + item de endividamento + colaborador).
            await session.execute(
                text(
                    "DELETE FROM payable_snapshots WHERE month = :m "
                    "AND (ref_id = :emp OR ref_id = :debt OR name LIKE :namelike)"
                ),
                {"m": pay, "emp": str(emp.id), "debt": str(debt.id), "namelike": f"Folha CLT {tag}%"},
            )
            await session.execute(
                text("DELETE FROM company_financial_items WHERE employee_id = :e"), {"e": str(emp.id)}
            )
            fresh = await session.get(Employee, emp.id)
            if fresh is not None:
                await session.delete(fresh)
            await session.commit()


if __name__ == "__main__":
    unittest.main()
