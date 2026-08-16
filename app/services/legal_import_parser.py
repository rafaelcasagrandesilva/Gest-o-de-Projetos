"""Transformação OFICIAL das fontes do Workspace Jurídico → payload normalizado.

Este módulo é a **única** implementação da transformação. Ele nasceu como o script
`scripts/build_legal_seed.py` (que gerou a carga da Fase 1) e foi extraído para cá quando a
planilha virou o formato oficial de alimentação do módulo: hoje o mesmo código serve ao
importador da tela, ao seed de desenvolvimento e ao script que regenera o JSON versionado.
Ninguém reimplementa regra de negócio — quem importa e quem semeia chamam esta função.

## As duas fontes, e por que só UMA é permanente

1. **Planilha unificada** (`.xlsx`, aba "Processos e Demitidos M&E") — é a BASE, é OBRIGATÓRIA e
   é a fonte PERMANENTE do módulo: pessoas (nome/CPF), empresa reclamada, contrato/projeto,
   datas de admissão/desligamento, status jurídico, acordo, rescisão, FGTS e observações.
2. **Painel de Passivo** (`painel_passivo.html`, bloco `const DATA`) — fonte **HISTÓRICA**,
   usada para ENRIQUECER a primeira carga: valor da causa, valor considerado, foro (tribunal),
   cidade, natureza, última movimentação e o link do JusBrasil.

Por que duas: a coluna "Valor Causa / Pedido (R$)" da planilha está 100% vazia, e "Valor
considerado", "Foro", "Última movimentação" e "Link JusBrasil" só existem no painel. As duas
fontes descrevem o MESMO conjunto de processos trabalhistas.

**O painel não é dependência do módulo.** Depois da carga inicial os metadados enriquecidos
pertencem ao BANCO, e as sincronizações seguintes usam apenas a planilha: `PANEL_ENRICHED_FIELDS`
lista esses campos e o serviço nunca os reescreve numa importação sem painel — nem com valor
vazio, nem com um substituto pior vindo da planilha (ver `legal_import_service`). Se o arquivo
do painel nunca mais existir, nada deixa de funcionar.

Chave de junção: nome normalizado do reclamante + tribunal deduzido do número CNJ
(segmento `.5.NN.` = TRT NN), com desempate por UF e, por fim, por ordem determinística.
Nenhuma entrada do painel é atribuída a dois processos (alocação consumível).

Fora de escopo: os processos cíveis/federais que existem só no painel (passivo da empresa, sem
vínculo com ex-colaborador). O modelo já os suporta (`case_type`); a carga segue a planilha.
"""

from __future__ import annotations

import io
import json
import re
import time
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Literal

import openpyxl

SHEET_NAME = "Processos e Demitidos M&E"
PAYLOAD_VERSION = 1

# --- colunas da planilha (índice 0-based) ---------------------------------------------------
C_NOME, C_CPF, C_POSSUI, C_NUM, C_EMPRESA, C_UF, C_PROJETO = 0, 1, 2, 3, 4, 5, 6
C_ADMISSAO, C_DESLIGAMENTO, C_AUDIENCIA, C_STATUS = 7, 8, 9, 10
C_VALOR_CAUSA, C_ACORDO, C_RESCISAO, C_PAGO, C_ABERTO, C_FGTS = 11, 12, 13, 14, 15, 16
C_OBS_JUR, C_OBS_RH = 22, 23

# Cabeçalho esperado — o contrato do "formato oficial". Só as colunas que a carga LÊ são
# verificadas (as demais podem mudar de nome sem quebrar a importação), e a comparação ignora
# acentos/caixa/espaços. Colunas extras ao final são aceitas: a planilha pode crescer.
EXPECTED_HEADER: dict[int, str] = {
    C_NOME: "Nome / Reclamante",
    C_CPF: "CPF",
    C_NUM: "Nº do Processo",
    C_EMPRESA: "Empresa Reclamada",
    C_UF: "UF / Local",
    C_PROJETO: "Contrato / Projeto",
    C_ADMISSAO: "Data de Admissão",
    C_DESLIGAMENTO: "Data de Desligamento",
    C_AUDIENCIA: "Data Audiência",
    C_STATUS: "Status do Processo (Jurídico)",
    C_VALOR_CAUSA: "Valor Causa / Pedido (R$)",
    C_ACORDO: "Valor Acordo Processo (R$)",
    C_RESCISAO: "Valor Rescisão (R$)",
    C_PAGO: "Valor Pago (R$)",
    C_ABERTO: "Valor em Aberto (R$)",
    C_FGTS: "Saldo FGTS (R$)",
    C_OBS_JUR: "Obs Jurídico",
    C_OBS_RH: "Obs RH",
}

# Status da planilha → enum `legal_case_status`. A planilha é a fonte oficial do status
# JURÍDICO (mais rica que a do painel, que é derivada da movimentação do JusBrasil).
STATUS_MAP: dict[str, str] = {
    "Em andamento": "EM_ANDAMENTO",
    "Com decisão/sentença": "COM_DECISAO",
    "Suspenso/Sobrestado": "SUSPENSO",
    "Encerrado/Arquivado": "ENCERRADO",
    "Acordo": "ACORDO",
    "Acordo Finalizado": "ACORDO_FINALIZADO",
    "Sem Processo Cadastrado": "SEM_PROCESSO",
}

# `classe` do painel → enum `legal_case_type`.
TYPE_MAP: dict[str, str] = {
    "trabalhista": "TRABALHISTA",
    "cível": "CIVEL",
    "federal": "TRIBUTARIO",
    "outro": "OUTRO",
}

# Entidades do grupo M&E: distinguem "Empresa" (nossa) de "Reclamado" (que pode ser a concessionária
# cliente, quando o ex-colaborador aciona também a tomadora do serviço).
GROUP_COMPANY_MARKERS = ("M&E", "M & E", "M E ENGENHARIA", "DSX")

# Projetos internos (área administrativa) não têm cliente externo.
INTERNAL_PROJECTS = {"Financeiro", "Administrativo", "Recursos Humanos"}

IssueLevel = Literal["ERROR", "WARNING"]

# Campos do PROCESSO cuja informação nasce no Painel de Passivo (fonte histórica). Numa
# importação sem o painel eles são deixados EM PAZ: o banco é a fonte oficial deles a partir da
# primeira carga. `defendant_name` está aqui por um motivo concreto — sem o painel ele cairia
# para a coluna "Empresa Reclamada" da planilha, que em ~25 processos registra outra parte;
# seria um valor PIOR sobrescrevendo um melhor, e não um campo vazio.
PANEL_ENRICHED_FIELDS: frozenset[str] = frozenset(
    {
        "jusbrasil_url",
        "amount_claimed",
        "amount_considered",
        "nature",
        "court",
        "city",
        "defendant_name",
        "last_movement",
        "last_movement_date",
    }
)


class LegalImportSourceError(ValueError):
    """Arquivo fora do formato oficial — a importação inteira é recusada."""


@dataclass
class SourceIssue:
    """Um problema encontrado na leitura. `ERROR` descarta a linha; `WARNING` só avisa."""

    level: IssueLevel
    message: str
    row: int | None = None
    identifier: str | None = None


@dataclass
class ParsedSources:
    payload: dict[str, Any]
    issues: list[SourceIssue] = field(default_factory=list)
    # Linhas que a planilha repete e a carga consolida num único registro (não é erro).
    duplicates: list[SourceIssue] = field(default_factory=list)
    # Linhas lidas e deliberadamente não carregadas (ex.: reclamante não identificado sem processo).
    skipped: list[SourceIssue] = field(default_factory=list)
    rows_read: int = 0
    panel_rows: int = 0
    panel_matched: int = 0
    elapsed_ms: int = 0

    @property
    def panel_present(self) -> bool:
        """O Painel de Passivo veio nesta carga? Governa a preservação dos campos enriquecidos."""
        return bool(self.payload.get("source", {}).get("panel"))


def norm_name(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", text).strip().upper()


def clean_str(value: object) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def parse_date(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_money(value: object) -> float | None:
    """Número simples ('7462.86') ou texto pt-BR ('1.234,56'). Retorna None quando vazio."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip().replace("R$", "").strip()
    if not text:
        return None
    text = text.replace(".", "").replace(",", ".") if "," in text else text
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def parse_agreement(text: str | None) -> float | None:
    """Soma o valor total de um acordo parcelado descrito em texto livre.

    A planilha registra o acordo como plano de pagamento, não como total:
    "3 X 2.333,34" → 7000.02 · "1 X 3250,00 e 4 X 2600,00" → 13650.00.
    Retorna None quando o texto não descreve nenhuma parcela reconhecível.
    """
    if not text:
        return None
    total = 0.0
    found = False
    for qty_raw, amount_raw in re.findall(r"(\d+)\s*[xX]\s*([\d.,]+)", text):
        amount = parse_money(amount_raw)
        if amount is None:
            continue
        total += int(qty_raw) * amount
        found = True
    if found:
        return round(total, 2)
    return parse_money(text)


# Códigos de tribunal da Justiça Estadual (segmento 8 do número CNJ) → sigla usada no painel.
STATE_COURTS: dict[int, str] = {
    1: "AC", 2: "AL", 3: "AP", 4: "AM", 5: "BA", 6: "CE", 7: "DF", 8: "ES", 9: "GO",
    10: "MA", 11: "MT", 12: "MS", 13: "MG", 14: "PA", 15: "PB", 16: "PR", 17: "PE",
    18: "PI", 19: "RJ", 20: "RN", 21: "RS", 22: "RO", 23: "RR", 24: "SC", 25: "SE",
    26: "SP", 27: "TO",
}

# Segmento do Judiciário (campo J do número CNJ) → enum `legal_case_type`.
SEGMENT_TYPE: dict[str, str] = {"5": "TRABALHISTA", "8": "CIVEL", "4": "TRIBUTARIO"}


def parse_cnj(number: str | None) -> tuple[str, int] | None:
    """(segmento, tribunal) do número CNJ (NNNNNNN-DD.AAAA.J.TR.OOOO).

    Tolerante à máscara de anonimização da fonte ("002XXXX-31.2026.5.24.0086") e a erros de
    digitação no prefixo: só os campos finais `.AAAA.J.TR.` importam.
    """
    if not number:
        return None
    match = re.search(r"\.(\d{4})\.(\d)\.(\d{2})\.", number)
    if not match:
        return None
    return match.group(2), int(match.group(3))


def court_from_case_number(number: str | None) -> str | None:
    """Sigla do tribunal a partir do número CNJ ("...5.24..." → TRT24; "...8.26..." → TJSP)."""
    parsed = parse_cnj(number)
    if not parsed:
        return None
    segment, court = parsed
    if segment == "5":
        return f"TRT{court}"
    if segment == "4":
        return f"TRF{court}"
    if segment == "8":
        uf = STATE_COURTS.get(court)
        return f"TJ{uf}" if uf else None
    return None


def type_from_case_number(number: str | None) -> str | None:
    """Tipo do processo pelo segmento do número CNJ (autoritativo quando o número é legível)."""
    parsed = parse_cnj(number)
    return SEGMENT_TYPE.get(parsed[0]) if parsed else None


def is_group_company(name: str | None) -> bool:
    """O nome identifica uma entidade do grupo M&E (e não a concessionária tomadora)?"""
    return bool(name) and any(m in str(name).upper() for m in GROUP_COMPANY_MARKERS)


def company_of(*candidates: str | None) -> str | None:
    """Empresa do processo: entidade do grupo M&E, se identificável; senão o 1º candidato.

    Muitas linhas trazem a concessionária (Energisa/Ampla/Eletropaulo) na coluna "Empresa
    Reclamada", porque o ex-colaborador aciona também a tomadora do serviço. Preferimos a
    entidade do grupo quando ela aparece; caso contrário mantemos exatamente o que a planilha
    registra — assim o filtro "Empresa" reflete o que o usuário já conhece, sem lacunas.
    """
    for candidate in candidates:
        if is_group_company(candidate):
            return candidate
    for candidate in candidates:
        if candidate:
            return candidate
    return None


def client_of(project: str | None) -> str | None:
    """Cliente deduzido do nome do contrato ("Energisa - C&M Naviraí" → "Energisa")."""
    if not project or project in INTERNAL_PROJECTS:
        return None
    return project.split(" - ")[0].strip() or None


def load_panel(content: bytes | str) -> list[dict[str, Any]]:
    """Extrai o bloco `const DATA` do Painel de Passivo."""
    html = content.decode("utf-8", errors="replace") if isinstance(content, bytes) else content
    match = re.search(r"^const DATA = (\[.*\]);$", html, re.M)
    if not match:
        raise LegalImportSourceError(
            "Bloco `const DATA` não encontrado no arquivo do Painel de Passivo. "
            "Envie o `painel_passivo.html` gerado pelo painel, sem edições."
        )
    try:
        return json.loads(match.group(1))
    except json.JSONDecodeError as exc:  # pragma: no cover - arquivo corrompido
        raise LegalImportSourceError(f"Painel de Passivo ilegível: {exc}") from exc


def allocate_panel_rows(
    sheet_rows: list[tuple], panel_rows: list[dict[str, Any]]
) -> dict[int, dict[str, Any]]:
    """Casa cada linha da planilha (com processo) a UMA entrada trabalhista do painel.

    Alocação consumível por nome: uma entrada do painel nunca é usada em dois processos. Dentro do
    mesmo nome, desempata por tribunal (deduzido do número CNJ), depois por UF, e por fim pela
    ordem determinística (maior valor primeiro) — os agregados por pessoa ficam corretos em
    qualquer caso, pois todos os processos pertencem à mesma pessoa.
    """
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entry in panel_rows:
        if entry.get("classe") == "trabalhista":
            by_name[norm_name(entry.get("rte"))].append(entry)
    for entries in by_name.values():
        entries.sort(key=lambda e: (-(e.get("valor") or 0), e.get("url") or ""))

    matched: dict[int, dict[str, Any]] = {}
    pending: list[tuple[int, tuple]] = []

    # 1ª passada: tribunal exato (mais específico).
    for index, row in sheet_rows:
        candidates = by_name.get(norm_name(row[C_NOME]))
        if not candidates:
            continue
        court = court_from_case_number(clean_str(row[C_NUM]))
        hits = [c for c in candidates if court and c.get("trib") == court]
        if len(hits) == 1:
            matched[index] = hits[0]
            candidates.remove(hits[0])
        else:
            pending.append((index, row))

    # 2ª passada: UF; 3ª: primeira restante.
    for index, row in pending:
        candidates = by_name.get(norm_name(row[C_NOME]))
        if not candidates:
            continue
        uf = clean_str(row[C_UF])
        hits = [c for c in candidates if uf and c.get("uf") == uf]
        chosen = hits[0] if hits else candidates[0]
        matched[index] = chosen
        candidates.remove(chosen)

    return matched


def _header_key(value: object) -> str:
    """Cabeçalho comparável, tolerante à digitação e intolerante ao arquivo errado.

    A verificação existe para barrar a planilha ERRADA, não para policiar tipografia: ignora
    acento, caixa, pontuação e as variações do indicador ordinal ('Nº', 'N°', 'N.', 'N' →
    'N DO PROCESSO'). Uma coluna com outro NOME, ou deslocada de posição, continua sendo recusada.
    """
    text = re.sub(r"[º°]", "", str(value or ""))
    return re.sub(r"\s+", " ", re.sub(r"[^0-9A-Za-z]+", " ", norm_name(text))).strip()


def _load_sheet(content: bytes | str) -> list[tuple]:
    """Abre a planilha, valida o formato oficial e devolve as linhas úteis (sem cabeçalho)."""
    source = io.BytesIO(content) if isinstance(content, bytes) else content
    try:
        workbook = openpyxl.load_workbook(source, data_only=True)
    except Exception as exc:  # openpyxl levanta tipos variados para arquivo inválido
        raise LegalImportSourceError(
            f"Não foi possível ler a planilha (.xlsx esperado): {exc}"
        ) from exc

    if SHEET_NAME not in workbook.sheetnames:
        raise LegalImportSourceError(
            f"Aba '{SHEET_NAME}' não encontrada. Abas do arquivo: "
            f"{', '.join(workbook.sheetnames) or '(nenhuma)'}."
        )

    all_rows = list(workbook[SHEET_NAME].iter_rows(values_only=True))
    if not all_rows:
        raise LegalImportSourceError(f"A aba '{SHEET_NAME}' está vazia.")

    header = all_rows[0]
    divergent = [
        f"coluna {index + 1}: esperado '{label}', encontrado '{clean_str(header[index]) or ''}'"
        for index, label in EXPECTED_HEADER.items()
        if index >= len(header) or _header_key(header[index]) != _header_key(label)
    ]
    if divergent:
        raise LegalImportSourceError(
            "Planilha fora do formato oficial do Jurídico — "
            + "; ".join(divergent[:4])
            + ("; …" if len(divergent) > 4 else "")
        )

    return [row for row in all_rows[1:] if any(cell is not None for cell in row)]


def _date_field(
    value: object, *, row_number: int, label: str, who: str | None, issues: list[SourceIssue]
) -> str | None:
    """Data com aviso: célula preenchida que não vira data é registrada, não silenciada."""
    parsed = parse_date(value)
    if parsed is None and clean_str(value):
        issues.append(
            SourceIssue(
                level="WARNING",
                row=row_number,
                identifier=who,
                message=f"{label} não reconhecida ('{clean_str(value)}') — campo ficou vazio.",
            )
        )
    return parsed


def build_payload(
    *,
    spreadsheet: bytes | str,
    panel: bytes | str | None = None,
    spreadsheet_name: str = "",
    panel_name: str | None = None,
) -> ParsedSources:
    """Lê as fontes e devolve o payload normalizado + o diagnóstico da leitura.

    Não toca no banco: é transformação pura, o que permite pré-visualizar a importação e gerar
    o JSON versionado com exatamente o mesmo código que a importação de produção executa.
    """
    started = time.perf_counter()
    issues: list[SourceIssue] = []
    duplicates: list[SourceIssue] = []
    skipped: list[SourceIssue] = []

    raw_rows = _load_sheet(spreadsheet)
    sheet_rows = list(enumerate(raw_rows))
    # Número da linha como o usuário vê no Excel (cabeçalho = linha 1).
    row_number = {index: index + 2 for index, _ in sheet_rows}

    panel_rows = load_panel(panel) if panel is not None else []
    panel_by_index = (
        allocate_panel_rows([(i, r) for i, r in sheet_rows if r[C_NUM]], panel_rows)
        if panel_rows
        else {}
    )

    # ---- pessoas: chave = CPF quando existe, senão nome normalizado -------------------------
    people: dict[str, dict[str, Any]] = {}
    person_key_by_row: dict[int, str] = {}

    for index, row in sheet_rows:
        line = row_number[index]
        name = clean_str(row[C_NOME])
        cpf = clean_str(row[C_CPF])
        number = clean_str(row[C_NUM])

        if not name:
            (issues if number else skipped).append(
                SourceIssue(
                    level="ERROR" if number else "WARNING",
                    row=line,
                    message=(
                        "Linha sem nome do reclamante."
                        if number
                        else "Linha sem nome e sem número de processo — ignorada."
                    ),
                )
            )
            continue

        # "Desconhecido" é um placeholder da planilha (reclamante não identificado), não uma
        # pessoa: o processo fica sem ex-colaborador vinculado (person_id nulo).
        if norm_name(name) == "DESCONHECIDO":
            skipped.append(
                SourceIssue(
                    level="WARNING",
                    row=line,
                    identifier=number or name,
                    message=(
                        "Reclamante não identificado — o processo entra sem pessoa vinculada."
                        if number
                        else "Reclamante não identificado e sem processo — linha ignorada."
                    ),
                )
            )
            continue

        key = f"cpf:{cpf}" if cpf else f"nome:{norm_name(name)}"
        person_key_by_row[index] = key

        empresa = clean_str(row[C_EMPRESA])
        projeto = clean_str(row[C_PROJETO])
        person = people.get(key)
        if person is None:
            people[key] = {
                "key": key,
                "full_name": name,
                "cpf": cpf,
                "company": company_of(empresa),
                "project": projeto,
                "client": client_of(projeto),
                "role": None,  # a planilha não traz cargo — campo já existe no modelo
                "admission_date": _date_field(
                    row[C_ADMISSAO], row_number=line, label="Data de admissão", who=name, issues=issues
                ),
                "termination_date": _date_field(
                    row[C_DESLIGAMENTO], row_number=line, label="Data de desligamento", who=name, issues=issues
                ),
                "severance_amount": parse_money(row[C_RESCISAO]),
                "fgts_balance": parse_money(row[C_FGTS]),
                "notes": clean_str(row[C_OBS_RH]),
            }
        else:
            duplicates.append(
                SourceIssue(
                    level="WARNING",
                    row=line,
                    identifier=name,
                    message="Pessoa repetida na planilha (uma linha por processo) — consolidada num único cadastro.",
                )
            )
            # Linhas repetidas do mesmo CPF (uma por processo): completa lacunas sem sobrescrever.
            for field_name, value in (
                ("company", company_of(empresa)),
                ("project", projeto),
                ("client", client_of(projeto)),
                ("admission_date", parse_date(row[C_ADMISSAO])),
                ("termination_date", parse_date(row[C_DESLIGAMENTO])),
                ("severance_amount", parse_money(row[C_RESCISAO])),
                ("fgts_balance", parse_money(row[C_FGTS])),
                ("notes", clean_str(row[C_OBS_RH])),
            ):
                if person.get(field_name) is None and value is not None:
                    person[field_name] = value

    # ---- processos --------------------------------------------------------------------------
    cases: list[dict[str, Any]] = []
    seen_numbers: dict[str, dict[str, Any]] = {}

    for index, row in sheet_rows:
        line = row_number[index]
        number = clean_str(row[C_NUM])
        if not number:
            continue  # ex-colaborador sem processo: vira só pessoa
        panel_entry = panel_by_index.get(index, {})
        empresa = clean_str(row[C_EMPRESA])
        projeto = clean_str(row[C_PROJETO])
        panel_defendant = clean_str(panel_entry.get("rdo"))
        status_label = clean_str(row[C_STATUS])
        agreement_terms = clean_str(row[C_ACORDO])
        claimant = clean_str(row[C_NOME])

        # Status desconhecido NÃO vira "Em andamento" à força: fica nulo, e o serviço aplica o
        # padrão só ao CRIAR. Assim uma reimportação não desfaz um status ajustado na tela.
        status = STATUS_MAP.get(status_label or "")
        if status_label and status is None:
            issues.append(
                SourceIssue(
                    level="WARNING",
                    row=line,
                    identifier=number,
                    message=f"Status '{status_label}' não reconhecido — status do processo não foi alterado.",
                )
            )

        if parse_cnj(number) is None:
            issues.append(
                SourceIssue(
                    level="WARNING",
                    row=line,
                    identifier=number,
                    message="Número fora do padrão CNJ — tribunal e tipo do processo não puderam ser deduzidos.",
                )
            )

        case = {
            "case_number": number,
            "person_key": person_key_by_row.get(index),
            "jusbrasil_url": clean_str(panel_entry.get("url")),
            "status": status,
            # Segmento do número CNJ é autoritativo; a classe do painel só cobre o que não é legível.
            "case_type": (
                type_from_case_number(number)
                or TYPE_MAP.get(clean_str(panel_entry.get("classe")) or "")
                or "TRABALHISTA"
            ),
            "nature": clean_str(panel_entry.get("natureza")),
            "uf": clean_str(row[C_UF]) or clean_str(panel_entry.get("uf")),
            "court": clean_str(panel_entry.get("trib")) or court_from_case_number(number),
            "city": clean_str(panel_entry.get("cidade")),
            "company": company_of(empresa, panel_defendant),
            "project": projeto,
            "client": client_of(projeto),
            "claimant_name": claimant,
            "defendant_name": panel_defendant or empresa,
            "amount_claimed": parse_money(panel_entry.get("valor")) or parse_money(row[C_VALOR_CAUSA]),
            "amount_considered": parse_money(panel_entry.get("vcons")),
            "agreement_terms": agreement_terms,
            "amount_agreed": parse_agreement(agreement_terms),
            "amount_paid": parse_money(row[C_PAGO]),
            "amount_pending": parse_money(row[C_ABERTO]),
            "last_movement": clean_str(panel_entry.get("umov")),
            "last_movement_date": parse_date(panel_entry.get("umdata")),
            "hearing_date": _date_field(
                row[C_AUDIENCIA], row_number=line, label="Data da audiência", who=number, issues=issues
            ),
            "distribution_date": None,  # ausente nas duas fontes
            "notes": clean_str(row[C_OBS_JUR]),
        }

        previous = seen_numbers.get(number)
        if previous is None:
            seen_numbers[number] = case
            cases.append(case)
            continue
        duplicates.append(
            SourceIssue(
                level="WARNING",
                row=line,
                identifier=number,
                message="Número de processo repetido na planilha — consolidado num único processo.",
            )
        )
        # Mesmo número em duas linhas: a planilha tem uma linha placeholder ("Desconhecido") e
        # outra com o reclamante identificado. Mantém UM processo, preferindo o identificado.
        if previous.get("person_key") is None and case.get("person_key") is not None:
            for field_name, value in case.items():
                if value is not None:
                    previous[field_name] = value

    for person in people.values():
        person["case_count_source"] = sum(1 for c in cases if c.get("person_key") == person["key"])

    payload = {
        "version": PAYLOAD_VERSION,
        "source": {
            "spreadsheet": spreadsheet_name,
            "sheet": SHEET_NAME,
            "panel": panel_name,
        },
        "people": sorted(people.values(), key=lambda p: norm_name(p["full_name"])),
        "cases": sorted(cases, key=lambda c: c["case_number"]),
    }

    return ParsedSources(
        payload=payload,
        issues=issues,
        duplicates=duplicates,
        skipped=skipped,
        rows_read=len(sheet_rows),
        panel_rows=len(panel_rows),
        panel_matched=len(panel_by_index),
        elapsed_ms=int((time.perf_counter() - started) * 1000),
    )


__all__ = [
    "LegalImportSourceError",
    "PANEL_ENRICHED_FIELDS",
    "ParsedSources",
    "SourceIssue",
    "SHEET_NAME",
    "build_payload",
    "load_panel",
]
