from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Sequence
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.services.export.builders import (
    build_pdf_bytes,
    build_projects_summary_pdf_bytes,
    build_projects_summary_xlsx_bytes,
    build_xlsx_bytes,
    format_brl,
    format_date_br,
)
from app.services.export.report_meta import ReportContext, friendly_filename, header_lines, report_title

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MIME_PDF = "application/pdf"


def _fname(report_type: str, ext: str, ctx: ReportContext | None) -> str:
    return friendly_filename(report_type, ext, periodo_token=ctx.periodo_token if ctx else None)


def _ident_meta(ctx: ReportContext | None, count: int | None = None) -> list[str]:
    """Identificação para o cabeçalho do PDF (título/'Gerado em' já vêm do builder).

    Usada apenas no PDF; o Excel abre direto nos dados, sem aba de identificação.
    """
    return header_lines(ctx, record_count=count, include_title=False, include_gen=False)

_MONTH_LABELS = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]


def _pct_br(ratio: float) -> str:
    return f"{ratio * 100:.2f}%".replace(".", ",")


def _month_keys(year: int) -> list[tuple[str, str]]:
    return [(lab, f"{year}-{i:02d}") for i, lab in enumerate(_MONTH_LABELS, start=1)]


def _payment_sum_month(pagamentos: list[dict], month_key: str) -> float:
    return sum(float(p.get("valor") or 0) for p in pagamentos if p.get("mes") == month_key)


def render_company_summary_bytes(
    data: dict[str, Any], fmt: str, ctx: ReportContext | None = None
) -> tuple[bytes, str, str]:
    summaries = data["rows"]
    competencia = data["competencia"]
    headers = [
        "Projeto",
        "Faturamento",
        "Folha",
        "Veículos",
        "Impostos",
        "Rateio",
        "Antecipação",
        "Retenção",
        "Lucro Operacional",
        "Lucro Líquido",
        "Margem %",
    ]
    xlsx_rows: list[list[Any]] = []
    pdf_rows: list[list[Any]] = []
    for r in summaries:
        xlsx_rows.append(
            [
                r["projeto"],
                r["faturamento"],
                r["folha"],
                r["veiculos"],
                r["impostos"],
                r["rateio"],
                r["antecipacao"],
                r["retencao"],
                r["lucro_operacional"],
                r["lucro_liquido"],
                r["margem"],
            ]
        )
        pdf_rows.append(
            [
                r["projeto"],
                format_brl(r["faturamento"]),
                format_brl(r["folha"]),
                format_brl(r["veiculos"]),
                format_brl(r["impostos"]),
                format_brl(r["rateio"]),
                format_brl(r["antecipacao"]),
                format_brl(r["retencao"]),
                format_brl(r["lucro_operacional"]),
                format_brl(r["lucro_liquido"]),
                _pct_br(float(r["margem"])),
            ]
        )
    if summaries:
        sf = sum(r["faturamento"] for r in summaries)
        t_folha = sum(r["folha"] for r in summaries)
        t_vei = sum(r["veiculos"] for r in summaries)
        t_imp = sum(r["impostos"] for r in summaries)
        t_rat = sum(r["rateio"] for r in summaries)
        t_ant = sum(r["antecipacao"] for r in summaries)
        t_ret = sum(r["retencao"] for r in summaries)
        t_lo = sum(r["lucro_operacional"] for r in summaries)
        t_nl = sum(r["lucro_liquido"] for r in summaries)
        t_marg = (t_nl / sf) if sf > 0 else 0.0
        totals_xlsx = ["TOTAL", sf, t_folha, t_vei, t_imp, t_rat, t_ant, t_ret, t_lo, t_nl, t_marg]
        totals_pdf = [
            "TOTAL",
            format_brl(sf),
            format_brl(t_folha),
            format_brl(t_vei),
            format_brl(t_imp),
            format_brl(t_rat),
            format_brl(t_ant),
            format_brl(t_ret),
            format_brl(t_lo),
            format_brl(t_nl),
            _pct_br(t_marg),
        ]
    else:
        totals_xlsx = ["TOTAL", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        totals_pdf = [
            "TOTAL",
            format_brl(0),
            format_brl(0),
            format_brl(0),
            format_brl(0),
            format_brl(0),
            format_brl(0),
            format_brl(0),
            format_brl(0),
            format_brl(0),
            _pct_br(0),
        ]
    title = report_title("company_summary")
    meta = _ident_meta(ctx, len(summaries)) + [
        f"Competência: {competencia}",
        f"Cenário: {data.get('scenario') or 'REALIZADO'}",
    ]
    if fmt == "xlsx":
        raw = build_projects_summary_xlsx_bytes(
            headers=headers, rows=xlsx_rows, totals_row=totals_xlsx
        )
        return raw, _fname("company_summary", "xlsx", ctx), MIME_XLSX
    raw = build_projects_summary_pdf_bytes(
        title=title, headers=headers, rows=pdf_rows, meta_lines=meta, totals_row=totals_pdf
    )
    return raw, _fname("company_summary", "pdf", ctx), MIME_PDF


def _summary_kpi_rows(s: dict[str, Any]) -> list[tuple[str, str]]:
    return [
        ("Faturamento", format_brl(float(s.get("total_revenue") or 0))),
        ("Folha (mão de obra)", format_brl(float(s.get("labor_cost") or 0))),
        ("Veículos", format_brl(float(s.get("vehicle_cost") or 0))),
        ("Sistemas", format_brl(float(s.get("system_cost") or 0))),
        ("Custos fixos operacionais", format_brl(float(s.get("fixed_operational_cost") or 0))),
        ("Impostos", format_brl(float(s.get("tax_amount") or 0))),
        ("Rateio", format_brl(float(s.get("overhead_amount") or 0))),
        ("Antecipação", format_brl(float(s.get("anticipation_amount") or 0))),
        ("Retenção", format_brl(float(s.get("total_retention") or 0))),
        ("Lucro operacional", format_brl(float(s.get("operational_profit") or 0))),
        ("Lucro líquido", format_brl(float(s.get("net_profit") or 0))),
        ("Margem líquida", _pct_br(float(s.get("margin_net") or 0))),
    ]


def render_project_summary_bytes(
    data: dict[str, Any], fmt: str, ctx: ReportContext | None = None
) -> tuple[bytes, str, str]:
    pname = data["project_name"]
    comp = data["competencia"][:7] if len(data["competencia"]) >= 7 else data["competencia"]
    s = data["summary"]

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumo"[:31]
        ws.cell(row=1, column=1, value=f"Cenário: {data.get('scenario') or 'REALIZADO'}")
        ws.cell(row=2, column=1, value="Indicador")
        ws.cell(row=2, column=2, value="Valor")
        ws.cell(row=2, column=1).font = Font(bold=True)
        ws.cell(row=2, column=2).font = Font(bold=True)
        for i, (k, v) in enumerate(_summary_kpi_rows(s), start=3):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v)
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 22

        def sheet_table(title: str, headers: list[str], rows: Sequence[Sequence[Any]]) -> None:
            nws = wb.create_sheet(title[:31])
            for c, h in enumerate(headers, start=1):
                cell = nws.cell(row=1, column=c, value=h)
                cell.font = Font(bold=True)
            for ri, row in enumerate(rows, start=2):
                for ci, val in enumerate(row, start=1):
                    nws.cell(row=ri, column=ci, value=val)
            for c in range(1, len(headers) + 1):
                nws.column_dimensions[get_column_letter(c)].width = 18

        labor = data["labor"]
        sheet_table(
            "Colaboradores",
            ["Nome", "Tipo", "% aloc.", "Custo alocado"],
            [
                (
                    x["name"],
                    x["tipo"],
                    f"{float(x['allocation_percentage']):.1f}%",
                    format_brl(float(x["allocated_cost"])),
                )
                for x in labor
            ],
        )
        veh = data["vehicles"]
        sheet_table(
            "Veículos",
            ["Placa", "Modelo", "Tipo", "Custo mês"],
            [
                (x["plate"], x.get("model") or "", x["vehicle_type"], format_brl(float(x["monthly_cost"])))
                for x in veh
            ],
        )
        sysrows = data["systems"]
        sheet_table(
            "Sistemas",
            ["Nome", "Valor"],
            [(x["name"], format_brl(float(x["value"]))) for x in sysrows],
        )
        fixrows = data["fixed_operational"]
        sheet_table(
            "Custos fixos",
            ["Nome", "Valor"],
            [(x["name"], format_brl(float(x["value"]))) for x in fixrows],
        )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue(), _fname("project_summary", "xlsx", ctx), MIME_XLSX

    # PDF
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = []
    story.append(Paragraph(f"<b>{escape('Projeto — resumo financeiro')}</b>", styles["Title"]))
    story.append(Paragraph(escape(f"{pname} · competência {comp}"), styles["Normal"]))
    story.append(
        Paragraph(escape(f"Cenário: {data.get('scenario') or 'REALIZADO'}"), styles["Normal"])
    )
    gen = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    story.append(Paragraph(f"Gerado em: {gen}", styles["Normal"]))
    for line in _ident_meta(ctx):
        story.append(Paragraph(escape(line), styles["Normal"]))
    story.append(Spacer(1, 0.3 * cm))

    kpi_data = [["Indicador", "Valor"]] + [[a, b] for a, b in _summary_kpi_rows(s)]
    tw = doc.width / 2
    t0 = Table(kpi_data, colWidths=[tw, tw])
    t0.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ]
        )
    )
    story.append(t0)
    story.append(Spacer(1, 0.4 * cm))

    def add_table(title: str, headers: list[str], rows: list[list[Any]]) -> None:
        story.append(Paragraph(f"<b>{escape(title)}</b>", styles["Heading2"]))
        d = [headers] + rows
        col_count = len(headers)
        w = doc.width / max(col_count, 1)
        tbl = Table(d, colWidths=[w] * col_count, repeatRows=1)
        tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ]
            )
        )
        story.append(tbl)
        story.append(Spacer(1, 0.35 * cm))

    labor = data["labor"]
    add_table(
        "Colaboradores",
        ["Nome", "Tipo", "%", "Custo alocado"],
        [
            [
                x["name"][:40],
                str(x["tipo"]),
                f"{float(x['allocation_percentage']):.1f}",
                format_brl(float(x["allocated_cost"])),
            ]
            for x in labor
        ],
    )
    veh = data["vehicles"]
    add_table(
        "Veículos",
        ["Placa", "Modelo", "Tipo", "Custo"],
        [
            [
                x["plate"],
                (x.get("model") or "")[:24],
                str(x["vehicle_type"]),
                format_brl(float(x["monthly_cost"])),
            ]
            for x in veh
        ],
    )
    sysrows = data["systems"]
    add_table(
        "Sistemas",
        ["Nome", "Valor"],
        [[x["name"][:50], format_brl(float(x["value"]))] for x in sysrows],
    )
    fixrows = data["fixed_operational"]
    add_table(
        "Custos fixos operacionais",
        ["Nome", "Valor"],
        [[x["name"][:50], format_brl(float(x["value"]))] for x in fixrows],
    )

    doc.build(story)
    buf.seek(0)
    return buf.getvalue(), _fname("project_summary", "pdf", ctx), MIME_PDF


def render_employees_bytes(
    data: dict[str, Any], fmt: str, ctx: ReportContext | None = None
) -> tuple[bytes, str, str]:
    cref = data["competencia_ref"]
    comp_display = cref[:10] if len(cref) >= 10 else cref
    headers = [
        "Nome", "E-mail", "Cargo", "Tipo", "Status",
        "Projetos", "Centros de custo",
        "Salário base", "Custos adicionais", "Custo total (cadastro)", "PJ horas/mês", "PJ custo adicional",
        "Periculosidade", "Adicional dirigida",
        "PIX tipo", "PIX chave",
        "Custo projetos (folha)", "Custo administrativo (folha)", "Custo total (folha)",
        "Criado em", "Atualizado em",
    ]
    rows = [
        [
            r["nome"], r.get("email") or "", r.get("cargo") or "", r["tipo"], r.get("status") or "",
            r.get("projetos") or "", r.get("centros_custo") or "",
            format_brl(r.get("salario_base")), format_brl(r.get("custos_adicionais")),
            format_brl(r.get("custo_total_cadastro")), r.get("pj_horas_mes", ""), format_brl(r.get("pj_custo_adicional")),
            r.get("periculosidade") or "", r.get("adicional_dirigida") or "",
            r.get("pix_tipo") or "", r.get("pix_chave") or "",
            format_brl(r.get("custo_projetos")), format_brl(r.get("custo_administrativo")), format_brl(r["custo"]),
            format_date_br(r["criado_em"]) if r.get("criado_em") else "",
            format_date_br(r["atualizado_em"]) if r.get("atualizado_em") else "",
        ]
        for r in data["rows"]
    ]
    meta = _ident_meta(ctx, len(rows)) + [
        f"Competência referência: {format_date_br(comp_display)}",
        f"Cenário: {data.get('scenario') or 'REALIZADO'}",
    ]
    title = report_title("employees")
    if fmt == "xlsx":
        raw = build_xlsx_bytes(
            headers=headers, rows=rows, sheet_title="Colaboradores"
        )
        return raw, _fname("employees", "xlsx", ctx), MIME_XLSX
    raw = build_pdf_bytes(title=title, headers=headers, rows=rows, meta_lines=meta)
    return raw, _fname("employees", "pdf", ctx), MIME_PDF


# Limite de caracteres do "Detalhamento dos Endividamentos" no PDF antes de resumir
# para "N endividamentos" (o espaço da célula é limitado). No XLSX mostra-se completo.
ENDIV_DETALHE_PDF_MAXLEN = 60


def _format_endividamentos_detalhe(itens: list[dict[str, Any]] | None, *, summarize: bool) -> str:
    """Detalhamento por dívida: "Descrição (R$ x); Descrição (R$ y)".

    - Um único item: sempre mostra ele por extenso.
    - `summarize` (PDF) com 2+ itens e texto longo: resume para "N endividamentos"
      (a coluna financeira "Endividamentos" continua mostrando o total).
    """
    itens = itens or []
    if not itens:
        return ""
    full = "; ".join(f"{it.get('descricao')} ({format_brl(it.get('valor'))})" for it in itens)
    if summarize and len(itens) >= 2 and len(full) > ENDIV_DETALHE_PDF_MAXLEN:
        return f"{len(itens)} endividamentos"
    return full


def render_payroll_bytes(
    data: dict[str, Any], fmt: str, ctx: ReportContext | None = None
) -> tuple[bytes, str, str]:
    """Folha de Pagamento: resumo executivo + 1 linha por colaborador + linha de totais.

    Novo relatório — NÃO altera o de Colaboradores (render_employees_bytes) nem os demais.
    """
    s = data.get("summary") or {}
    summarize_detalhe = fmt != "xlsx"  # PDF resume quando faltar espaço; XLSX completo.
    headers = [
        "Nome", "E-mail", "Cargo", "Tipo", "Status",
        "Centro de Custo", "Distribuição",
        "Salário Base", "Salário Líquido / Contratado", "Benefícios", "VR",
        "VT", "Ajuda de custo", "Endividamentos", "Detalhamento dos Endividamentos",
        "Outros pagamentos",
        "Total da Folha", "PIX Tipo", "PIX Chave",
    ]
    rows: list[list[Any]] = []
    for r in data["rows"]:
        rows.append(
            [
                r["nome"], r.get("email") or "", r.get("cargo") or "", r["tipo"], r.get("status") or "",
                r.get("centro_custo") or "—",
                r.get("distribuicao") or "—",
                format_brl(r.get("salario_base")),
                format_brl(r.get("salario_liquido")),
                format_brl(r.get("beneficios")),
                format_brl(r.get("vr")),
                format_brl(r.get("vt")),
                format_brl(r.get("ajuda_custo")),
                format_brl(r.get("endividamentos")),
                _format_endividamentos_detalhe(r.get("endividamentos_itens"), summarize=summarize_detalhe),
                format_brl(r.get("outros")),
                format_brl(r.get("total_folha")),
                r.get("pix_tipo") or "",
                r.get("pix_chave") or "",
            ]
        )

    # Resumo executivo (antes da tabela — igual aos demais relatórios).
    meta = _ident_meta(ctx, len(rows)) + [
        f"Competência: {s.get('competencia') or data.get('competencia_ref')}",
        f"Cenário: {data.get('scenario') or 'REALIZADO'}",
        f"Colaboradores CLT: {s.get('qtd_clt', 0)}; PJ: {s.get('qtd_pj', 0)}",
        f"Total Salários: {format_brl(s.get('total_salarios'))}",
        f"Total Benefícios: {format_brl(s.get('total_beneficios'))}",
        f"Total VR: {format_brl(s.get('total_vr'))}",
        f"Total Endividamentos: {format_brl(s.get('total_endividamentos'))}",
        f"Total Ajuda de Custo: {format_brl(s.get('total_ajuda_custo'))}",
        f"TOTAL GERAL DA FOLHA: {format_brl(s.get('total_geral'))}",
    ]

    # Linha de totais (posicional, alinhada às colunas de valor).
    total_base = sum(float(r.get("salario_base") or 0) for r in data["rows"])
    totals_row = [
        "TOTAIS", "", "", "", "", "", "",
        format_brl(total_base),
        format_brl(s.get("total_salarios")),
        format_brl(s.get("total_beneficios")),
        format_brl(s.get("total_vr")),
        "",  # VT (futuro)
        format_brl(s.get("total_ajuda_custo")),
        format_brl(s.get("total_endividamentos")),
        "",  # Detalhamento dos Endividamentos (sem total)
        "",  # Outros (futuro)
        format_brl(s.get("total_geral")),
        "", "",
    ]

    title = report_title("payroll")
    if fmt == "xlsx":
        raw = build_xlsx_bytes(
            headers=headers, rows=rows, sheet_title="Folha de Pagamento", totals_row=totals_row
        )
        return raw, _fname("payroll", "xlsx", ctx), MIME_XLSX
    raw = build_pdf_bytes(title=title, headers=headers, rows=rows, meta_lines=meta, totals_row=totals_row)
    return raw, _fname("payroll", "pdf", ctx), MIME_PDF


def render_vehicles_bytes(
    data: dict[str, Any], fmt: str, ctx: ReportContext | None = None
) -> tuple[bytes, str, str]:
    headers = ["Placa", "Modelo", "Descrição", "Tipo", "Centro de Custo", "Condutor", "Custo mensal", "Status", "Criado em", "Atualizado em"]
    rows = [
        [
            r["placa"],
            r.get("modelo") or "",
            r.get("descricao") or "",
            r["tipo"],
            r.get("centro_custo") or "—",
            r.get("condutor") or "—",
            format_brl(r["custo_mensal"]),
            "Ativo" if r["ativo"] else "Inativo",
            format_date_br(r["criado_em"]) if r.get("criado_em") else "",
            format_date_br(r["atualizado_em"]) if r.get("atualizado_em") else "",
        ]
        for r in data["rows"]
    ]
    meta = _ident_meta(ctx, len(rows)) + [f"Somente ativos: {'sim' if data['active_only'] else 'não'}"]
    title = report_title("vehicles")
    if fmt == "xlsx":
        raw = build_xlsx_bytes(
            headers=headers, rows=rows, sheet_title="Frota"
        )
        return raw, _fname("vehicles", "xlsx", ctx), MIME_XLSX
    raw = build_pdf_bytes(title=title, headers=headers, rows=rows, meta_lines=meta)
    return raw, _fname("vehicles", "pdf", ctx), MIME_PDF


def render_invoices_bytes(
    data: dict[str, Any], fmt: str, ctx: ReportContext | None = None
) -> tuple[bytes, str, str]:
    headers = ["Projeto", "Nº NF", "Valor bruto", "Vencimento", "Recebido", "Saldo", "Status"]
    rows = []
    sum_bruto = sum_rec = sum_saldo = 0.0
    for r in data["rows"]:
        vb = float(r["valor_bruto"])
        tr = float(r["total_recebido"])
        sd = float(r["saldo"])
        sum_bruto += vb
        sum_rec += tr
        sum_saldo += sd
        rows.append(
            [
                r["projeto"],
                r["numero_nf"],
                format_brl(vb),
                format_date_br(r["vencimento"]),
                format_brl(tr),
                format_brl(sd),
                r["status"],
            ]
        )
    meta = _ident_meta(ctx, len(rows))
    totals = ["Totais", "", format_brl(sum_bruto), "", format_brl(sum_rec), format_brl(sum_saldo), ""]
    title = report_title("invoices")
    if fmt == "xlsx":
        raw = build_xlsx_bytes(
            headers=headers, rows=rows, sheet_title="NFs", totals_row=totals
        )
        return raw, _fname("invoices", "xlsx", ctx), MIME_XLSX
    raw = build_pdf_bytes(title=title, headers=headers, rows=rows, meta_lines=meta, totals_row=totals)
    return raw, _fname("invoices", "pdf", ctx), MIME_PDF


def render_company_finance_matrix_bytes(
    data: dict[str, Any], fmt: str, ctx: ReportContext | None = None
) -> tuple[bytes, str, str]:
    tipo = data["tipo"]
    competencia = data["competencia"]
    year = int(str(competencia)[:4])
    items = data["items"]
    month_pairs = _month_keys(year)
    headers = ["Item", "Referência"] + [lab for lab, _ in month_pairs] + ["Total pago", "Saldo / Restante", "Progresso %"]
    is_debt = tipo == "endividamento"
    label = "Endividamento" if is_debt else "Custos fixos (empresa)"
    xlsx_rows: list[list[Any]] = []
    pdf_rows: list[list[Any]] = []

    for it in items:
        pags = it.get("pagamentos") or []
        ref = float(it.get("valor_referencia") or 0)
        row_x: list[Any] = [it.get("nome") or "", ref]
        row_p: list[Any] = [it.get("nome") or "", format_brl(ref)]
        for _, mk in month_pairs:
            v = _payment_sum_month(pags, mk)
            row_x.append(v)
            row_p.append(format_brl(v))
        tp = float(it.get("total_pago") or 0)
        row_x.append(tp)
        row_p.append(format_brl(tp))
        rest = it.get("restante")
        if is_debt:
            row_x.append(float(rest) if rest is not None else "")
            row_p.append("—" if rest is None else format_brl(float(rest)))
        else:
            row_x.append("")
            row_p.append("—")
        prog = float(it.get("progresso") or 0)
        row_x.append(prog)
        row_p.append(f"{prog * 100:.1f}%".replace(".", ","))
        xlsx_rows.append(row_x)
        pdf_rows.append(row_p)

    report_type = "debt" if is_debt else "fixed_costs"
    meta = _ident_meta(ctx, len(items)) + [f"Competência ref.: {competencia}; ano das colunas: {year}"]
    title = report_title(report_type)

    if fmt == "xlsx":
        raw = build_xlsx_bytes(
            headers=headers, rows=xlsx_rows, sheet_title=label[:31]
        )
        return raw, _fname(report_type, "xlsx", ctx), MIME_XLSX
    raw = build_pdf_bytes(title=title, headers=headers, rows=pdf_rows, meta_lines=meta)
    return raw, _fname(report_type, "pdf", ctx), MIME_PDF


def render_users_bytes(
    data: dict[str, Any], fmt: str, ctx: ReportContext | None = None
) -> tuple[bytes, str, str]:
    headers = ["E-mail", "Nome", "Ativo", "Papéis", "Criado em", "Atualizado em"]
    rows = [
        [
            r["email"],
            r["nome"],
            "Sim" if r["ativo"] else "Não",
            r["papeis"],
            format_date_br(r["criado_em"]) if r.get("criado_em") else "",
            format_date_br(r["atualizado_em"]) if r.get("atualizado_em") else "",
        ]
        for r in data["rows"]
    ]
    meta = _ident_meta(ctx, len(rows))
    title = report_title("users")
    if fmt == "xlsx":
        raw = build_xlsx_bytes(
            headers=headers, rows=rows, sheet_title="Usuários"
        )
        return raw, _fname("users", "xlsx", ctx), MIME_XLSX
    raw = build_pdf_bytes(title=title, headers=headers, rows=rows, meta_lines=meta)
    return raw, _fname("users", "pdf", ctx), MIME_PDF


def render_revenues_bytes(
    data: dict[str, Any], fmt: str, ctx: ReportContext | None = None
) -> tuple[bytes, str, str]:
    headers = ["Projeto", "Competência", "Cenário", "Valor", "Descrição", "Status", "Retenção", "Criado em", "Atualizado em"]
    rows = []
    for r in data["rows"]:
        c = r["competencia"]
        cd = c[:10] if len(c) >= 10 else c
        rows.append(
            [
                r.get("projeto") or r.get("project_id") or "",
                format_date_br(cd),
                r.get("cenario") or "",
                format_brl(r["valor"]),
                (r["descricao"] or "")[:200],
                r["status"],
                "Sim" if r["retencao"] else "Não",
                format_date_br(r["criado_em"]) if r.get("criado_em") else "",
                format_date_br(r["atualizado_em"]) if r.get("atualizado_em") else "",
            ]
        )
    meta = _ident_meta(ctx, len(rows))
    title = report_title("revenues")
    if fmt == "xlsx":
        raw = build_xlsx_bytes(
            headers=headers, rows=rows, sheet_title="Receitas"
        )
        return raw, _fname("revenues", "xlsx", ctx), MIME_XLSX
    raw = build_pdf_bytes(title=title, headers=headers, rows=rows, meta_lines=meta)
    return raw, _fname("revenues", "pdf", ctx), MIME_PDF


def render_dashboard_bytes(
    data: dict[str, Any], fmt: str, ctx: ReportContext | None = None
) -> tuple[bytes, str, str]:
    summary = data["summary"]["summary"]
    series = data["summary"]["monthly_series"]
    months = data["months"]
    headers = [
        "Mês",
        "Receita",
        "Custo total",
        "Custo operacional total",
        "Mão de obra",
        "Veículos",
        "Sistemas",
        "Fixos operacionais",
        "Imposto",
        "Rateio",
        "Antecipação",
        "Retenção",
        "Lucro operacional",
        "Lucro líquido",
        "EBITDA",
        "Margem líquida %",
    ]
    data_rows = []
    for m in series:
        data_rows.append(
            [
                format_date_br(m["competencia"]),
                format_brl(float(m["total_revenue"])),
                format_brl(float(m["total_cost"])),
                format_brl(float(m.get("operational_cost") or 0)),
                format_brl(float(m.get("labor_cost") or 0)),
                format_brl(float(m.get("vehicle_cost") or 0)),
                format_brl(float(m.get("system_cost") or 0)),
                format_brl(float(m.get("fixed_operational_cost") or 0)),
                format_brl(float(m.get("tax_amount") or 0)),
                format_brl(float(m.get("overhead_amount") or 0)),
                format_brl(float(m.get("anticipation_amount") or 0)),
                format_brl(float(m.get("total_retention") or 0)),
                format_brl(float(m.get("operational_profit") or 0)),
                format_brl(float(m["net_profit"])),
                format_brl(float(m.get("ebitda") or 0)),
                f"{float(m['margin_net']) * 100:.1f}%",
            ]
        )
    sm = summary
    scen = (data.get("summary") or {}).get("scenario") or "REALIZADO"
    meta = _ident_meta(ctx, len(data_rows)) + [
        f"Competência ref.: {format_date_br(sm['competencia'])}",
        f"Cenário (série principal): {scen}",
        f"Série: últimos {months} meses",
        f"Receita: {format_brl(float(sm['total_revenue']))}; Custo: {format_brl(float(sm['total_cost']))}; "
        f"Lucro líq.: {format_brl(float(sm['net_profit']))}",
    ]
    title = report_title("dashboard")
    if fmt == "xlsx":
        raw = build_xlsx_bytes(
            headers=headers, rows=data_rows, sheet_title="Dashboard"
        )
        return raw, _fname("dashboard", "xlsx", ctx), MIME_XLSX
    raw = build_pdf_bytes(title=title, headers=headers, rows=data_rows, meta_lines=meta, totals_row=None)
    return raw, _fname("dashboard", "pdf", ctx), MIME_PDF


def render_report_bytes(
    report_type: str, payload: dict[str, Any], fmt: str, ctx: ReportContext | None = None
) -> tuple[bytes, str, str]:
    if fmt not in ("xlsx", "pdf"):
        raise ValueError("formato inválido")
    if report_type == "project_summary":
        return render_project_summary_bytes(payload, fmt, ctx)
    if report_type == "company_summary":
        return render_company_summary_bytes(payload, fmt, ctx)
    if report_type == "employees":
        return render_employees_bytes(payload, fmt, ctx)
    if report_type == "payroll":
        return render_payroll_bytes(payload, fmt, ctx)
    if report_type == "vehicles":
        return render_vehicles_bytes(payload, fmt, ctx)
    if report_type == "invoices":
        return render_invoices_bytes(payload, fmt, ctx)
    if report_type in ("debt", "fixed_costs"):
        return render_company_finance_matrix_bytes(payload, fmt, ctx)
    if report_type == "dashboard":
        return render_dashboard_bytes(payload, fmt, ctx)
    if report_type == "users":
        return render_users_bytes(payload, fmt, ctx)
    if report_type == "revenues":
        return render_revenues_bytes(payload, fmt, ctx)
    raise ValueError(f"tipo de relatório desconhecido: {report_type}")
