from __future__ import annotations

import io
from datetime import date, datetime, timezone
from typing import Any, Sequence
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def format_brl(n: float | int | None) -> str:
    if n is None:
        return ""
    try:
        v = float(n)
    except (TypeError, ValueError):
        return str(n)
    return _brl_py(v)


def _brl_py(v: float) -> str:
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def format_date_br(d: date | datetime | str | None) -> str:
    if d is None:
        return ""
    if isinstance(d, str):
        if len(d) >= 10 and d[4] == "-":
            y, m, dd = d[:10].split("-")
            return f"{dd}/{m}/{y}"
        return d
    if isinstance(d, datetime):
        d = d.date()
    return f"{d.day:02d}/{d.month:02d}/{d.year}"


def build_xlsx_bytes(
    *,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    sheet_title: str = "Exportação",
    totals_row: Sequence[Any] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=str(h))
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if totals_row is not None:
        r = len(rows) + 3
        for c_idx, val in enumerate(totals_row, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_pdf_bytes(
    *,
    title: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    meta_lines: Sequence[str],
    totals_row: Sequence[Any] | None = None,
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = []
    story.append(Paragraph(f"<b>{escape(title)}</b>", styles["Title"]))
    gen = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    story.append(Paragraph(f"Gerado em: {gen}", styles["Normal"]))
    for line in meta_lines:
        if line:
            story.append(Paragraph(escape(line), styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))

    data: list[list[Any]] = [list(headers)]
    for row in rows:
        data.append([("" if v is None else str(v)) for v in row])
    if totals_row is not None:
        data.append([("" if v is None else str(v)) for v in totals_row])

    col_count = max(len(r) for r in data) if data else 1
    w = doc.width / max(col_count, 1)
    tbl = Table(data, colWidths=[w] * col_count, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    if totals_row is not None:
        last = len(data) - 1
        tbl.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, last), (-1, last), "Helvetica-Bold"),
                    ("BACKGROUND", (0, last), (-1, last), colors.HexColor("#f8fafc")),
                ]
            )
        )
    story.append(tbl)
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def export_filename(module_slug: str, ext: str, period_suffix: str | None = None) -> str:
    suf = period_suffix or datetime.now(timezone.utc).strftime("%Y-%m")
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in module_slug)
    return f"{safe}_{suf}.{ext}"


_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_BRL_NUM_FMT = '[$R$-416] #,##0.00'
_PCT_NUM_FMT = "0.00%"


def _apply_negative_red(cell) -> None:
    v = cell.value
    if isinstance(v, (int, float)) and v < 0:
        cell.font = Font(color="FF0000")


def build_projects_summary_xlsx_bytes(
    *,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    totals_row: Sequence[Any],
) -> bytes:
    """Planilha executiva: cabeçalho amarelo, totais amarelos, negativos em vermelho; valores numéricos em moeda."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo por projeto"[:31]
    ncols = len(headers)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=str(h))
        cell.font = Font(bold=True)
        cell.fill = _YELLOW
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    margin_col = ncols
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c_idx == 1:
                continue
            if c_idx == margin_col:
                if isinstance(val, (int, float)):
                    cell.number_format = _PCT_NUM_FMT
                    cell.value = float(val)
                _apply_negative_red(cell)
            else:
                if isinstance(val, (int, float)):
                    cell.number_format = _BRL_NUM_FMT
                    cell.value = float(val)
                _apply_negative_red(cell)
    tr = len(rows) + 2
    for c_idx, val in enumerate(totals_row, start=1):
        cell = ws.cell(row=tr, column=c_idx, value=val)
        is_neg = isinstance(val, (int, float)) and val < 0
        cell.font = Font(bold=True, color="FF0000") if is_neg else Font(bold=True)
        cell.fill = _YELLOW
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if c_idx == 1:
            continue
        if c_idx == margin_col:
            if isinstance(val, (int, float)):
                cell.number_format = _PCT_NUM_FMT
                cell.value = float(val)
        else:
            if isinstance(val, (int, float)):
                cell.number_format = _BRL_NUM_FMT
                cell.value = float(val)
    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_projects_summary_pdf_bytes(
    *,
    title: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    meta_lines: Sequence[str],
    totals_row: Sequence[Any],
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = []
    story.append(Paragraph(f"<b>{escape(title)}</b>", styles["Title"]))
    gen = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    story.append(Paragraph(f"Gerado em: {gen}", styles["Normal"]))
    for line in meta_lines:
        if line:
            story.append(Paragraph(escape(line), styles["Normal"]))
    story.append(Spacer(1, 0.35 * cm))

    data: list[list[Any]] = [list(headers)]
    for row in rows:
        data.append([("" if v is None else str(v)) for v in row])
    data.append([("" if v is None else str(v)) for v in totals_row])

    col_count = max(len(r) for r in data) if data else 1
    w = doc.width / max(col_count, 1)
    tbl = Table(data, colWidths=[w] * col_count, repeatRows=1)
    yellow = colors.HexColor("#FFFF00")
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), yellow),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    last = len(data) - 1
    tbl.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, last), (-1, last), "Helvetica-Bold"),
                ("BACKGROUND", (0, last), (-1, last), yellow),
            ]
        )
    )
    story.append(tbl)
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()
