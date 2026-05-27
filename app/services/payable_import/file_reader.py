from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any

from openpyxl import load_workbook

from app.services.payable_import.constants import MAX_IMPORT_BYTES, MAX_IMPORT_ROWS
from app.services.payable_import.normalization import cell_str


@dataclass(frozen=True)
class RawSheet:
    """Linhas brutas indexadas a partir de 1 (como no Excel)."""

    rows: list[tuple[int, tuple[Any, ...]]]


@dataclass(frozen=True)
class SpreadsheetTable:
    header_row: int
    columns: list[str]
    rows: list[tuple[int, dict[str, Any]]]


def _detect_file_kind(filename: str, content: bytes) -> str:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return "csv"
    if name.endswith(".xlsx"):
        return "xlsx"
    if content[:2] == b"PK":
        return "xlsx"
    return "csv"


def read_raw_sheet(*, content: bytes, filename: str = "") -> RawSheet:
    if len(content) > MAX_IMPORT_BYTES:
        raise ValueError("Arquivo muito grande (máximo 5 MB).")
    kind = _detect_file_kind(filename, content)
    if kind == "xlsx":
        return _read_xlsx_raw(content)
    return _read_csv_raw(content)


def _read_xlsx_raw(content: bytes) -> RawSheet:
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Arquivo Excel inválido ou corrompido.") from exc
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("Planilha vazia.")
    out: list[tuple[int, tuple[Any, ...]]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        out.append((idx, row))
        if idx > MAX_IMPORT_ROWS + 50:
            break
    wb.close()
    if not out:
        raise ValueError("Planilha vazia.")
    return RawSheet(rows=out)


def _read_csv_raw(content: bytes) -> RawSheet:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            text = None
    if text is None:
        raise ValueError("Não foi possível ler o CSV (encoding inválido).")
    lines = [l for l in text.splitlines() if l.strip()]
    if not lines:
        raise ValueError("Arquivo CSV vazio.")

    def _count_outside_quotes(s: str, ch: str) -> int:
        count = 0
        in_quotes = False
        i = 0
        while i < len(s):
            c = s[i]
            if c == '"':
                # Em CSV, aspas podem aparecer duplicadas como escape: "".
                # Mantemos uma regra simples para não contar delimitadores dentro de campos.
                in_quotes = not in_quotes
            elif c == ch and not in_quotes:
                count += 1
            i += 1
        return count

    first_line = lines[0]
    sc = _count_outside_quotes(first_line, ";")
    cc = _count_outside_quotes(first_line, ",")
    delimiter = ";" if sc > cc else ","

    reader = csv.reader(StringIO(text), delimiter=delimiter)
    out: list[tuple[int, tuple[Any, ...]]] = []
    for idx, row in enumerate(reader, start=1):
        out.append((idx, tuple(row)))
        if idx > MAX_IMPORT_ROWS + 50:
            break
    if not out:
        raise ValueError("Arquivo CSV vazio.")
    return RawSheet(rows=out)


def _build_columns(header_cells: tuple[Any, ...]) -> list[str]:
    columns: list[str] = []
    seen: dict[str, int] = {}
    for cell in header_cells:
        label = cell_str(cell)
        if not label:
            columns.append("")
            continue
        base = label
        n = seen.get(base, 0)
        if n:
            label = f"{base} ({n + 1})"
        seen[base] = n + 1
        columns.append(label)
    return columns


def table_from_raw(*, raw: RawSheet, header_row: int) -> SpreadsheetTable:
    if header_row < 1:
        raise ValueError("Linha do cabeçalho deve ser >= 1.")
    header: tuple[Any, ...] | None = None
    for line_no, cells in raw.rows:
        if line_no == header_row:
            header = cells
            break
    if header is None:
        raise ValueError(f"Cabeçalho não encontrado na linha {header_row}.")

    columns = _build_columns(header)
    active_columns = [c for c in columns if c]
    if not active_columns:
        raise ValueError("Nenhuma coluna identificada no cabeçalho.")

    col_index = {name: i for i, name in enumerate(columns) if name}
    data_rows: list[tuple[int, dict[str, Any]]] = []
    data_count = 0
    for line_no, cells in raw.rows:
        if line_no <= header_row:
            continue
        row_dict: dict[str, Any] = {}
        for name, idx in col_index.items():
            if idx < len(cells):
                row_dict[name] = cells[idx]
        if not any(cell_str(v) for v in row_dict.values()):
            continue
        data_rows.append((line_no, row_dict))
        data_count += 1
        if data_count > MAX_IMPORT_ROWS:
            raise ValueError(f"Planilha excede o limite de {MAX_IMPORT_ROWS} linhas de dados.")

    return SpreadsheetTable(header_row=header_row, columns=active_columns, rows=data_rows)


def _serialize_sample_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return ""
    return cell_str(value) if not isinstance(value, (int, float)) else value


def sample_rows(table: SpreadsheetTable, *, limit: int) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for _line_no, row in table.rows[:limit]:
        out.append({k: _serialize_sample_value(v) for k, v in row.items()})
    return out
